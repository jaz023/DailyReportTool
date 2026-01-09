from __future__ import annotations
import os, glob
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

TEMPLATE_XLSX = "template.xlsx"
MAPPING_XLSX  = "mapping.xlsx"
SOURCE_GLOB   = "sources/*.csv"
OUTPUT_DIR    = "output"

# 你想指定也可以，但就算寫錯也會自動找最合理的
TARGET_SHEET_NAME = "daily rev0(+cn)"

COL_TIME  = "Time"
COL_NAME  = "Name"
COL_VALUE = "Value"

# 找不到資料時要填什麼（可改成 "-"）
FILL_IF_MISSING = "NA"

# ====== ✅ 把時間/日期寫進報表（你可自行改 cell）======
WRITE_DATETIME_TO_TEMPLATE = True
DATE_CELL = "E34"
TIME_CELL = "E35"
DATETIME_CELL = None  # e.g. "B2" 如果你想寫成 "2025-12-25 01:33" 一格

# ====== ✅ 診斷開關 ======
DEBUG_MAPPING = True
DEBUG_MISS_CANDIDATES = True
MISS_CANDIDATE_TOPK = 5


def ensure_dir(p: str) -> None:
    os.makedirs(p, exist_ok=True)


def parse_center_time(t_str: str) -> datetime:
    """
    必須手動輸入（不可空白）：
    - YYYY-MM-DD HH:MM
    - YYYY/MM/DD HH:MM（順手支援）
    """
    s = (t_str or "").strip()
    if not s:
        raise ValueError("時間不可空白。請輸入 YYYY-MM-DD HH:MM（例如 2025-12-25 01:35）。")

    for fmt in ("%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass

    raise ValueError("時間格式錯誤。請輸入 YYYY-MM-DD HH:MM（例如 2025-12-25 01:35）。")


def agg_values(df: pd.DataFrame, agg: str) -> float | None:
    if df.empty:
        return None
    agg = (agg or "LAST").upper().strip()

    df2 = df.copy()
    df2[COL_VALUE] = pd.to_numeric(df2[COL_VALUE], errors="coerce")
    df2 = df2.dropna(subset=[COL_VALUE, COL_TIME])
    if df2.empty:
        return None

    if agg == "LAST":
        df2 = df2.sort_values(COL_TIME)
        return float(df2.iloc[-1][COL_VALUE])
    if agg == "AVG":
        return float(df2[COL_VALUE].mean())
    if agg == "MAX":
        return float(df2[COL_VALUE].max())
    if agg == "MIN":
        return float(df2[COL_VALUE].min())
    if agg == "SUM":
        return float(df2[COL_VALUE].sum())

    df2 = df2.sort_values(COL_TIME)
    return float(df2.iloc[-1][COL_VALUE])


def match_rows(df: pd.DataFrame, source_name: str) -> pd.DataFrame:
    """
    名稱模糊比對：
    1) 完全相等
    2) Name 包含 source_name（不分大小寫）
    3) source_name 用空白切開做 AND（每個關鍵字都必須出現）
    """
    name = str(source_name).strip()
    if not name or df.empty:
        return df.iloc[0:0]

    # 保證可以做 .str.contains（避免非字串型態）
    name_series = df[COL_NAME].astype(str)

    # 1) 完全相等
    exact = df[name_series == name]
    if not exact.empty:
        return exact

    # 2) 包含整段字串
    contains = df[name_series.str.contains(name, case=False, na=False)]
    if not contains.empty:
        return contains

    # 3) 多關鍵字 AND
    keywords = [k for k in name.split() if k]
    mask = pd.Series(True, index=df.index)
    for kw in keywords:
        mask &= name_series.str.contains(kw, case=False, na=False)

    return df[mask]


def miss_candidates(df_win: pd.DataFrame, src: str, topk: int = 5) -> list[str]:
    """
    針對 mapping 的 SourceName（src），在 df_win 的 Name 欄位中找可能的候選（簡易 scoring）：
    - 把 src 以空白切關鍵字
    - 候選 Name 中包含越多關鍵字，分數越高
    """
    if df_win.empty:
        return []

    src = str(src).strip()
    if not src:
        return []

    candidates = df_win[COL_NAME].dropna().astype(str).unique().tolist()
    kws = [k for k in src.split() if k]
    if not kws:
        return []

    hits = []
    for n in candidates:
        n_low = n.lower()
        score = sum(1 for k in kws if k.lower() in n_low)
        if score > 0:
            hits.append((score, n))

    hits.sort(key=lambda x: (-x[0], x[1]))
    return [h[1] for h in hits[:topk]]


def read_sources() -> pd.DataFrame:
    files = sorted(glob.glob(SOURCE_GLOB))
    if not files:
        raise FileNotFoundError("sources 資料夾中找不到 CSV 檔案")

    dfs = []

    for fp in files:
        print(f"讀取資料檔：{fp}")

        df = pd.read_csv(
            fp,
            skiprows=6,          # 跳過 metadata
            encoding="utf-8-sig"
        )

        print("COLUMNS:", df.columns.tolist())

        SRC_TIME_COL = "time"
        if SRC_TIME_COL not in df.columns:
            print(f"警告：{fp} 找不到 time 欄位，跳過")
            continue

        value_columns = [c for c in df.columns if c not in ("no.", SRC_TIME_COL)]
        if not value_columns:
            print(f"警告：{fp} 沒有可用的數值欄位，跳過")
            continue

        df[SRC_TIME_COL] = pd.to_datetime(df[SRC_TIME_COL], errors="coerce")

        df_long = df.melt(
            id_vars=[SRC_TIME_COL],
            value_vars=value_columns,
            var_name=COL_NAME,
            value_name=COL_VALUE
        )

        # 統一欄名成 Time / Name / Value
        df_long = df_long.rename(columns={SRC_TIME_COL: COL_TIME})
        df_long[COL_NAME] = df_long[COL_NAME].astype(str).str.strip()

        # 時間一定要有效
        df_long = df_long.dropna(subset=[COL_TIME])

        if not df_long.empty:
            dfs.append(df_long)

    if not dfs:
        raise ValueError("所有 CSV 在指定格式下都沒有有效資料")

    return pd.concat(dfs, ignore_index=True)


def pick_sheet(wb, preferred: str):
    """依照偏好名稱（忽略大小寫/空白）挑 sheet；挑不到就選含 +CN 的；再不行才報錯。"""
    def norm(s: str) -> str:
        return "".join(str(s).lower().split())

    preferred_norm = norm(preferred)
    sheet_map = {norm(n): n for n in wb.sheetnames}

    # 1) 正規化精準匹配
    if preferred_norm in sheet_map:
        return wb[sheet_map[preferred_norm]], sheet_map[preferred_norm]

    # 2) 退而求其次：找包含 CN 的 sheet
    cn_candidates = [n for n in wb.sheetnames if "CN" in n.upper()]
    if cn_candidates:
        return wb[cn_candidates[0]], cn_candidates[0]

    raise ValueError(f"模板找不到合適工作表。你指定：{preferred}；現有：{wb.sheetnames}")


def _safe_write_cell(ws, addr: str, value) -> None:
    """
    安全寫入：如果 addr 落在 merged cell 的非左上角，openpyxl 會是 MergedCell（read-only）。
    這裡會自動改寫到該 merged range 的左上角 cell。
    """
    cell = ws[addr]
    if isinstance(cell, MergedCell):
        for r in ws.merged_cells.ranges:
            if addr in r:
                ws.cell(row=r.min_row, column=r.min_col).value = value
                return
        # 理論上不會到這裡；若到了，保守不寫
        return
    cell.value = value


def write_datetime(ws, center_time: datetime, start_time: datetime, end_time: datetime) -> None:
    """
    把日期/時間/時間窗寫進表格。
    - 若你在上方設定 DATE_CELL / TIME_CELL / DATETIME_CELL，就寫到那些格子
    - 若未設定，預設寫到 A1~A3（不覆蓋原本有內容的格子）
    """
    if not WRITE_DATETIME_TO_TEMPLATE:
        return

    date_str = center_time.strftime("%Y-%m-%d")
    time_str = center_time.strftime("%H:%M")
    dt_str   = center_time.strftime("%Y-%m-%d %H:%M")
    win_str  = f"{start_time.strftime('%Y-%m-%d %H:%M:%S')} ~ {end_time.strftime('%Y-%m-%d %H:%M:%S')}"

    # 1) 使用者指定 cell（含 merged cell 安全寫入）
    if DATETIME_CELL:
        _safe_write_cell(ws, DATETIME_CELL, dt_str)
    if DATE_CELL:
        _safe_write_cell(ws, DATE_CELL, date_str)
    if TIME_CELL:
        _safe_write_cell(ws, TIME_CELL, time_str)

    # 2) 若沒有指定，寫到 A1~A3（但避免覆蓋已經有值的格子）
    if not any([DATETIME_CELL, DATE_CELL, TIME_CELL]):
        if ws["A1"].value in (None, ""):
            ws["A1"].value = f"Report Time: {dt_str}"
        if ws["A2"].value in (None, ""):
            ws["A2"].value = f"Window: {win_str}"
        if ws["A3"].value in (None, ""):
            ws["A3"].value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"


def main():
    # ====== ✅ 手動輸入（不可空白），且打錯會要求重輸入 ======
    while True:
        try:
            t_str = input("請輸入要填表的時間 (YYYY-MM-DD HH:MM): ").strip()
            center_time = parse_center_time(t_str)
            break
        except ValueError as e:
            print("[ERROR]", e)

    while True:
        try:
            minutes = int(input("請輸入抓取區間(分鐘，表示前後各幾分鐘，例如 30): ").strip())
            if minutes < 0:
                raise ValueError("分鐘不可為負數。")
            break
        except Exception as e:
            print("[ERROR] 分鐘輸入錯誤，請輸入整數，例如 30。")

    start_time = center_time - timedelta(minutes=minutes)
    end_time   = center_time + timedelta(minutes=minutes)

    print(f"抓取範圍：{start_time} ~ {end_time}")

    map_df = pd.read_excel(MAPPING_XLSX, sheet_name="Map")
    for col in ["SourceName", "TargetCell"]:
        if col not in map_df.columns:
            raise ValueError(f"mapping.xlsx 的 Map 必須包含欄位：{col}")
    if "Agg" not in map_df.columns:
        map_df["Agg"] = "LAST"

    map_df["SourceName"] = map_df["SourceName"].astype(str).str.strip()
    map_df["TargetCell"] = map_df["TargetCell"].astype(str).str.strip()
    map_df["Agg"] = map_df["Agg"].astype(str).str.strip()

    # ====== ✅ Mapping 診斷：rows + sample ======
    if DEBUG_MAPPING:
        print("\n[DEBUG] Mapping rows:", len(map_df))
        try:
            print("[DEBUG] Mapping head:")
            print(map_df[["SourceName", "TargetCell", "Agg"]].head(10).to_string(index=False))
        except Exception:
            print("[DEBUG] Mapping columns:", map_df.columns.tolist())

    data = read_sources()

    # 只取時間窗內的資料
    df_win = data[(data[COL_TIME] >= start_time) & (data[COL_TIME] <= end_time)].copy()

    if df_win.empty:
        print("\n[WARN] 時間窗內 df_win 為空，請加大 minutes 或確認輸入時間是否在資料範圍內。")

    wb = load_workbook(TEMPLATE_XLSX)

    ws, real_sheet = pick_sheet(wb, TARGET_SHEET_NAME)
    print("使用工作表：", real_sheet)

    # ====== ✅ 寫入報表時間資訊（手動輸入的 center_time） ======
    write_datetime(ws, center_time, start_time, end_time)

    filled = 0
    missing_items = []

    for _, row in map_df.iterrows():
        src = row["SourceName"]
        addr = row["TargetCell"]
        agg = row["Agg"]

        sub = match_rows(df_win, src)
        v = agg_values(sub, agg)

        if v is None:
            _safe_write_cell(ws, addr, FILL_IF_MISSING)
            missing_items.append(src)

            # ====== ✅ MISS candidates 診斷 ======
            if DEBUG_MISS_CANDIDATES:
                cands = miss_candidates(df_win, src, topk=MISS_CANDIDATE_TOPK)
                if cands:
                    print(f"[MISS] {src} -> candidates: {cands}")
                else:
                    print(f"[MISS] {src} -> no candidates")
            continue

        _safe_write_cell(ws, addr, v)
        filled += 1

    ensure_dir(OUTPUT_DIR)
    out_path = os.path.join(
        OUTPUT_DIR,
        f"DailyReport_{center_time.strftime('%Y%m%d_%H%M')}_pm{minutes}.xlsx"
    )
    wb.save(out_path)

    print(f"\n完成：寫入 {filled} 項")
    if missing_items:
        print(f"以下品項在此時間範圍找不到資料（已填入 {FILL_IF_MISSING}，請確認名稱或時間範圍）：")
        for x in missing_items[:30]:
            print(" -", x)
        if len(missing_items) > 30:
            print(f" ...（共 {len(missing_items)} 項）")

    print("\n輸出檔：", out_path)


if __name__ == "__main__":
    main()
